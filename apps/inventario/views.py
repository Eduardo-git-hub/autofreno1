from django.shortcuts import render,redirect
from django.core.exceptions import ObjectDoesNotExist
from django.views.generic import TemplateView,ListView,UpdateView,CreateView,DeleteView
from django.forms import formset_factory
from django.views.generic.edit import FormView
from django.urls import reverse_lazy
from django.http.response import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment,Border,Font,PatternFill,Side
from .models import Proveedor,Pastilla,Producto
from .forms import ProveedorForm,PastillaForm

# Create your views here.

class Inicio(TemplateView):
    template_name = 'index.html'

class ListarProveedor(ListView):
    model = Proveedor
    template_name = 'inventario/proveedor/listar_proveedor.html'
    queryset = Proveedor.objects.all()
    paginate_by = 12

    def get_queryset(self, *args, **kwargs):
        qs = super().get_queryset(*args, **kwargs)
        bnombre = self.request.GET.get('filtrarn')
        if bnombre:
            return qs.filter(nombre__icontains = bnombre)
        return qs

class CrearProveedor(CreateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'inventario/proveedor/crear_proveedor.html'
    success_url = reverse_lazy('inventario:listar_proveedor')

class ActualizarProveedor(UpdateView):
    model = Proveedor
    form_class = ProveedorForm
    template_name = 'inventario/proveedor/crear_proveedor.html'
    success_url = reverse_lazy('inventario:listar_proveedor')

class EliminarProveedor(DeleteView):
    model = Proveedor
    success_url = reverse_lazy('inventario:listar_proveedor')

class ListarPastillas(ListView):
    model = Pastilla
    template_name = 'inventario/pastilla/listar_pastilla.html'
    paginate_by = 12

    def get_queryset(self, *args, **kwargs):
        qs = super().get_queryset(*args, **kwargs)
        bmarca = self.request.GET.get('filtrarma')
        b_año = self.request.GET.get('filtraño')
        bubicacion = self.request.GET.get('filtrarR')
        bcarro = self.request.GET.get('filtrarcr')
        if bmarca:
            return qs.filter(marca__icontains = bmarca)
        if b_año:
            return qs.filter(años__icontains = b_año)
        if bubicacion:
            return qs.filter(ubicacion__icontains = bubicacion)
        if bcarro:
            return qs.filter(carro__icontains = bcarro)
        return qs

class CrearPastillas(CreateView):
    model = Pastilla
    form_class = PastillaForm
    template_name = 'inventario/pastilla/crear_pastilla.html'
    success_url = reverse_lazy('inventario:listar_pastilla')

class ActualizarPastillas(UpdateView):
    model = Pastilla
    form_class = PastillaForm
    template_name = 'inventario/pastilla/crear_pastilla.html'
    success_url = reverse_lazy('inventario:listar_pastilla')

class EliminarPastilla(DeleteView):
    model = Producto
    success_url = reverse_lazy('inventario:listar_pastilla')

class Reportes_P(ListView):
    model = Pastilla
    template_name = 'inventario/reportes/reportes_productos.html'
    paginate_by = 12


class Report_P(TemplateView):
    def get(self,request,*args,**kwargs):
        fub = request.GET.get('filtrarc')
        fab = request.GET.get('ubicar')
        fmarc = request.GET.get('filmarc')
        productos = Pastilla.objects.filter(codigo__contains = fub, ubicacion__contains = fab, marca__contains = fmarc)
        wb = Workbook()
        ws = wb.active
        cont = 1
        ws.title = 'Hoja'+str(cont)

        ws.protection.enable()
        #Crear el titulo de la hoja
        ws['B2'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B2'].border = Border(left= Side(border_style= "thin"), right= Side(border_style="thin"),
                                 top= Side(border_style="thin"), bottom= Side(border_style="thin"))
        ws['B2'].font = Font(name="Time New Roman", size= 12, bold= True)
        ws['B2'] = 'REPORTE DE PASTILLAS'

        ws.merge_cells('B2:F2')

        #Cambiar las caracteristicas de las celdas
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20

        #Crear la cabecera
        ws['B4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['B4'].border = Border(left= Side(border_style="medium"), right= Side(border_style="medium"),
                                 top= Side(border_style="medium"), bottom= Side(border_style="medium"))
        ws['B4'].font = Font(name="Time New Roman", size= 10, bold= True)
        ws['B4'] = 'Codigo'
        ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['C4'].border = Border(left= Side(border_style="medium"), right= Side(border_style="medium"),
                                 top= Side(border_style="medium"), bottom= Side(border_style="medium"))
        ws['C4'].font = Font(name="Time New Roman", size= 10, bold= True)
        ws['C4'] = 'Cantidad'
        ws['D4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['D4'].border = Border(left= Side(border_style="medium"), right= Side(border_style="medium"),
                                 top= Side(border_style="medium"), bottom= Side(border_style="medium"))
        ws['D4'].font = Font(name="Time New Roman", size= 10, bold= True)
        ws['D4'] = 'Marca'
        ws['E4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['E4'].border = Border(left= Side(border_style="medium"), right= Side(border_style="medium"),
                                 top= Side(border_style="medium"), bottom= Side(border_style="medium"))
        ws['E4'].font = Font(name="Time New Roman", size= 10, bold= True)
        ws['E4'] = 'Repuesto'
        ws['F4'].alignment = Alignment(horizontal="center", vertical="center")
        ws['F4'].border = Border(left= Side(border_style="medium"), right= Side(border_style="medium"),
                                 top= Side(border_style="medium"), bottom= Side(border_style="medium"))
        ws['F4'].font = Font(name="Time New Roman", size= 10, bold= True)
        ws['F4'] = 'Ubicación'


        contador = 5

        for q in productos:
            ws.cell(row = contador, column = 2).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = contador, column = 2).border = Border(left= Side(border_style="thin"), right= Side(border_style="thin"),
                                    top= Side(border_style="thin"), bottom= Side(border_style="thin"))
            ws.cell(row = contador, column = 2).value = q.codigo
            ws.cell(row = contador, column = 3).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = contador, column = 3).border = Border(left= Side(border_style="thin"), right= Side(border_style="thin"),
                                    top= Side(border_style="thin"), bottom= Side(border_style="thin"))
            ws.cell(row = contador, column = 3).value = q.cantidad #str(q.provedor_codigo.all())
            ws.cell(row = contador, column = 4).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = contador, column = 4).border = Border(left= Side(border_style="thin"), right= Side(border_style="thin"),
                                    top= Side(border_style="thin"), bottom= Side(border_style="thin"))
            ws.cell(row = contador, column = 4).value = q.marca
            ws.cell(row = contador, column = 5).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = contador, column = 5).border = Border(left= Side(border_style="thin"), right= Side(border_style="thin"),
                                    top= Side(border_style="thin"), bottom= Side(border_style="thin"))
            ws.cell(row = contador, column = 5).value = q.repuesto
            ws.cell(row = contador, column = 6).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row = contador, column = 6).border = Border(left= Side(border_style="thin"), right= Side(border_style="thin"),
                                    top= Side(border_style="thin"), bottom= Side(border_style="thin"))
            ws.cell(row = contador, column = 6).value = q.ubicacion
            contador += 1


        nombre_archivo = "Reporte_Productos.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
